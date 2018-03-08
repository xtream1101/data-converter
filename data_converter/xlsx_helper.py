from openpyxl import load_workbook
from pyexcelerate import Workbook
from data_converter import utils

# ============== #
# Read Functions #
# ============== #


def _list_from_xlsx_with_header(wb, sheet_name):
    """
    Read a sheet from an xlsx file, and convert data to dicts in
    which the keys are the header values

    Args:
        wb (workbook object): openpyxl workbook object
        sheet_name (string): Sheet in the xlsx file to read

    Returns:
        (list): A list of dicts from data in the xlsx sheet
    """

    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb[wb.sheetnames[0]]

    try:
        # Python 3
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    except AttributeError:
        # Python 2.7
        headers = [cell.value for cell in sheet.iter_rows(min_row=1, max_row=1).next()]

    sheet_width = len(headers)
    sheet_iterator = sheet.iter_rows(min_row=2)

    # iterate through each sheet_iteartor, and append the headers and row value of index i, from 0 to
    # to width of the sheet (number of columns)
    sheet_data = [
        {headers[i]: row[i].value for i in range(sheet_width)}
        for row in sheet_iterator
    ]

    return sheet_data


def _list_from_xlsx_without_header(wb, sheet_name):
    """
       Read a sheet from an xlsx file, and convert data to lists in

       Args:
           wb (workbook object): openpyxl workbook object
           sheet_name (string): Sheet in the xlsx file to read

       Returns:
            (list): A list of lists from data in the xlsx sheet
       """
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb[wb.sheetnames[0]]

    sheet_iterator = sheet.iter_rows(min_row=1)
    # iterate through each sheet_iteartor, and append the headers and row value of index i, from 0 to
    # to width of the sheet (number of columns)
    sheet_data = [
        [cell.value for cell in row]
        for row in sheet_iterator
    ]

    return sheet_data


def read_file(input_file, sheet_name=None, header=True, **kwargs):
    """
    Read a file and return a list of a list of dicts if header is true,
    where the key values are the header values, or a list of lists if
    header is true. If no sheet name is given, will read the first
    sheet of the given xlsx file

    Args:
        input_file (string): Location of file to be read
        sheet_name (string): Sheet name to read in xlsx file
        header (boolean): whether or not the xlsx file has headers
        **kwargs: Arbitrary keyword arguments
    Returns:
         (list): Of either dicts or lists of data from the xlsx sheet
    """
    # data_only returns values in xlsx files rather than formulae
    ext = utils._get_file_ext(input_file)
    if ext != 'xlsx':
        raise ValueError("Expecting a xlsx file, but got: {ext}".format(ext=ext))

    wb = load_workbook(filename=input_file, data_only=True, read_only=True)
    sheet_data = _list_from_xlsx_with_header(wb, sheet_name) if header else _list_from_xlsx_without_header(wb, sheet_name)

    return sheet_data


# =============== #
# Write Functions #
# =============== #


def write_file(output_data, output_file, has_header=True, **kwargs):
    """Write a list of lists/dics to an excel file

    Args:
        output_data (list): list of lists/dicts of data to be saved
        output_file (str): Path to xlsx file

    Named Args:
        **kwargs: Catches all other args that were passed but do not care about

    Returns:
        str: Absolute file path of the saved file

    """
    output_file = utils._get_absolute_path(output_file)

    # if it's a list of lists, just write it to the workbook
    if isinstance(output_data[0], (list, tuple)):
        wb = Workbook()
        wb.new_sheet("Sheet1", data=output_data)
        wb.save(output_file)

    # if it's a list of dicts more logic
    else:
        # convert list of dicts to list of lists
        output = [d.values() for d in output_data]

        # if they want a header, create it from the dict keys
        if has_header:
            output[0:0] = [output_data[0].keys()]

        wb = Workbook()
        wb.new_sheet("Sheet1", data=output)
        wb.save(output_file)

    return output_file
